"""
╔══════════════════════════════════════════════════════════════╗
║   QUANT RADAR BOT v4.1 — Fix Edition                         ║
║                                                              ║
║   NAPRAWIONE vs v4.0:                                        ║
║   ✅ Elastyczne dopasowanie kluczy z batch API               ║
║   ✅ Poprawne screener/exchange dla każdej giełdy            ║
║   ✅ Obsługa DU=FWB, XC=Euronext, HEL=Finland               ║
║   ✅ Kolumny auto-fit we WSZYSTKICH arkuszach Excel          ║
║   ✅ Tryb --sprawdz do testowania pojedynczego tickera       ║
║   ✅ Raport "co nie działa i dlaczego" w konsoli             ║
╚══════════════════════════════════════════════════════════════╝
"""

import os, json, time, argparse, warnings
from datetime import datetime
import pandas as pd
import requests

warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════════════════════
#   PRZYCZYNY DLACZEGO TICKERY MOGĄ NIE DZIAŁAĆ
#   (wyjaśnienie dla każdego przypadku z raportu)
# ══════════════════════════════════════════════════════════════
#
#  KGH.WA   → alias KGH ≠ KGHM w TV. Naprawka: dodany alias
#  AMD.DE   → TV zwraca klucz "FWB:AMD" lub "XETR:AMD" zamiast "XETRA:AMD"
#             Naprawka: elastyczne dopasowanie klucza po samym symbolu
#  AIR.DE   → Airbus na XETRA. TV może zwrócić "XETR:AIR" (skrót!)
#  RHM.DE   → Rheinmetall. Poprawny, problem był w dopasowaniu klucza
#  INGA.AS  → ING Groep. Screener netherlands/AMS - poprawny
#  NESTE.HE → Neste. Screener finland/HEL - poprawny
#  W9C.DU   → DU=Düsseldorf. TV używa screener=germany, exchange=FWB (Frankfurt)
#             Większość spółek z DU dostępna też na XETRA lub FWB
#  AMSE.XC  → XC=Euronext. TV screener="euronext", exchange="EURONEXT"
#             Przykład: AMSE to Euronext Brussels/Amsterdam
#  UFG      → United Fire Group. Sprawdź na TV czy nadal notowany
#  12DA.DE  → Może być niedostępny na TV lub pod inną nazwą
#  ERCG.DE  → Sprawdź dokładną nazwę symbolu na tradingview.com
#  SPL.WA   → Sprawdź czy SPL istnieje na GPW (może być mała/wycofana spółka)
#  ATJ.WA   → Sprawdź czy ATJ istnieje na GPW

# ══════════════════════════════════════════════════════════════
#   MAPA INSTRUMENTÓW
#   Format: "ticker_w_portfelu": ("screener_TV", "EXCHANGE:SYMBOL")
#
#   UWAGA: Exchange musi być dokładnie taki jak TV używa wewnętrznie
#   Sprawdź na: tradingview.com → szukaj spółki → URL zawiera exchange
# ══════════════════════════════════════════════════════════════

MAPA: dict[str, tuple[str, str]] = {

    # ── GPW Polska ──────────────────────────────────────────
    "PKO.WA":   ("poland",      "GPW:PKO"),
    "KGHM.WA":  ("poland",      "GPW:KGHM"),
    "KGH.WA":   ("poland",      "GPW:KGHM"),   # alias KGH = KGHM na GPW
    "PKN.WA":   ("poland",      "GPW:PKN"),
    "PZU.WA":   ("poland",      "GPW:PZU"),
    "CDR.WA":   ("poland",      "GPW:CDR"),
    "LPP.WA":   ("poland",      "GPW:LPP"),
    "ALE.WA":   ("poland",      "GPW:ALE"),
    "DNP.WA":   ("poland",      "GPW:DNP"),
    "DIG.WA":   ("poland",      "GPW:DIG"),
    "MBK.WA":   ("poland",      "GPW:MBK"),
    "OPL.WA":   ("poland",      "GPW:OPL"),
    "ALR.WA":   ("poland",      "GPW:ALR"),
    "ASB.WA":   ("poland",      "GPW:ASB"),
    "ELT.WA":   ("poland",      "GPW:ELT"),
    "SPL.WA":   ("poland",      "GPW:SPL"),
    "ATJ.WA":   ("poland",      "GPW:ATJ"),
    "PCO.WA":   ("poland",      "GPW:PCO"),
    "PKP.WA":   ("poland",      "GPW:PKP"),
    "TPE.WA":   ("poland",      "GPW:TPE"),
    "TEN.WA":   ("poland",      "GPW:TEN"),

    # ── Niemcy XETRA ────────────────────────────────────────
    # UWAGA: TV często skraca XETRA → XETR w zwracanych kluczach!
    # Dopasowanie po samym symbolu (np. "AMD") omija ten problem
    "AMD.DE":   ("germany",     "XETRA:AMD"),
    "AIR.DE":   ("germany",     "XETRA:AIR"),    # Airbus SE
    "RHM.DE":   ("germany",     "XETRA:RHM"),    # Rheinmetall
    "CBK.DE":   ("germany",     "XETRA:CBK"),    # Commerzbank
    "ERCG.DE":  ("germany",     "XETRA:ERCG"),   # sprawdź na TV
    "12DA.DE":  ("germany",     "XETRA:12DA"),   # może niedostępny
    "PTX.DE":   ("germany",     "XETRA:PTX"),
    "SAP.DE":   ("germany",     "XETRA:SAP"),
    "ALV.DE":   ("germany",     "XETRA:ALV"),    # Allianz

    # ── Niemcy FWB (Frankfurt/Düsseldorf) ───────────────────
    # DU = Düsseldorf → TV używa exchange=FWB dla tych papierów
    "W9C.DU":   ("germany",     "FWB:W9C"),

    # ── Holandia Amsterdam ──────────────────────────────────
    "INGA.AS":  ("netherlands", "AMS:INGA"),     # ING Groep
    "ASML.AS":  ("netherlands", "AMS:ASML"),

    # ── Euronext (XC = Euronext Brussels/Amsterdam) ─────────
    # Błąd w v4.0: XC mapowało na NASDAQ! Poprawka: screener=euronext
    "AMSE.XC":  ("euronext",    "EURONEXT:AMSE"),

    # ── Finlandia Helsinki ──────────────────────────────────
    "NESTE.HE": ("finland",     "HEL:NESTE"),
    "NOKIA.HE": ("finland",     "HEL:NOKIA"),

    # ── US NASDAQ ───────────────────────────────────────────
    "AAPL":     ("america",     "NASDAQ:AAPL"),
    "MSFT":     ("america",     "NASDAQ:MSFT"),
    "NVDA":     ("america",     "NASDAQ:NVDA"),
    "GOOGL":    ("america",     "NASDAQ:GOOGL"),
    "AMZN":     ("america",     "NASDAQ:AMZN"),
    "META":     ("america",     "NASDAQ:META"),
    "TSLA":     ("america",     "NASDAQ:TSLA"),
    "AMD":      ("america",     "NASDAQ:AMD"),
    "ASML":     ("america",     "NASDAQ:ASML"),
    "COKE":     ("america",     "NASDAQ:COKE"),
    "QQQ":      ("america",     "NASDAQ:QQQ"),
    "SOXX":     ("america",     "NASDAQ:SOXX"),

    # ── US NYSE / AMEX ──────────────────────────────────────
    "HAL":      ("america",     "NYSE:HAL"),
    "NVO":      ("america",     "NYSE:NVO"),
    "APH":      ("america",     "NYSE:APH"),
    "UFG":      ("america",     "NYSE:UFG"),     # sprawdź czy nadal notowany
    "GOLD":     ("america",     "NYSE:GOLD"),    # Barrick Gold
    "SPY":      ("america",     "AMEX:SPY"),
    "VTI":      ("america",     "AMEX:VTI"),
    "IWM":      ("america",     "AMEX:IWM"),
    "XLK":      ("america",     "AMEX:XLK"),

    # ── Forex / Surowce ─────────────────────────────────────
    "EURUSD=X": ("forex",       "FX:EURUSD"),
    "USDPLN=X": ("forex",       "FX:USDPLN"),
    "GBPUSD=X": ("forex",       "FX:GBPUSD"),
    "XAUUSD":   ("forex",       "FX:XAUUSD"),
    "XAGUSD":   ("forex",       "FX:XAGUSD"),
}

# Auto-wykrywanie dla nieznanych tickerów (sufiks → screener/exchange)
SUFIKS_AUTO: dict[str, tuple[str, str]] = {
    "WA": ("poland",      "GPW"),
    "DE": ("germany",     "XETRA"),
    "DU": ("germany",     "FWB"),      # Düsseldorf → FWB w TV
    "AS": ("netherlands", "AMS"),
    "HE": ("finland",     "HEL"),
    "LS": ("portugal",    "EURONEXT"),
    "PA": ("france",      "EURONEXT"),
    "L":  ("uk",          "LSE"),
    "SW": ("switzerland", "SIX"),
    "XC": ("euronext",    "EURONEXT"), # ← BUG FIX: było NASDAQ!
    "MI": ("italy",       "MIL"),
    "MC": ("spain",       "BME"),
}

def tv_info(ticker: str) -> tuple[str, str] | None:
    """Zwraca (screener, 'EXCHANGE:SYMBOL') lub None."""
    if ticker in MAPA:
        return MAPA[ticker]
    if "." in ticker:
        czesc, suf = ticker.rsplit(".", 1)
        if suf in SUFIKS_AUTO:
            scr, exc = SUFIKS_AUTO[suf]
            return (scr, f"{exc}:{czesc}")
    if "=" not in ticker and "/" not in ticker:
        return ("america", f"NASDAQ:{ticker}")
    return None


# ══════════════════════════════════════════════════════════════
PLIK_PORTFELA    = "invest_grupy.json"
DOMYSLNE_GRUPY   = {"💼 MÓJ PORTFEL": ["PKO.WA", "AAPL", "KGHM.WA"]}
TELEGRAM_TOKEN   = os.environ.get("TELEGRAM_TOKEN",   "8725724228:AAHj-xQtVV87YQyE7ZQfdCgGyM6ICCX2520").strip(" '\"")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "7510513100").strip(" '\"")

WAGI = {"1h": 15, "4h": 25, "1d": 40, "1wk": 20}
TV_INTERWAL = {
    "1h":  "INTERVAL_1_HOUR",
    "4h":  "INTERVAL_4_HOURS",
    "1d":  "INTERVAL_1_DAY",
    "1wk": "INTERVAL_1_WEEK",
}

def zaladuj_portfel() -> dict:
    if not os.path.exists(PLIK_PORTFELA):
        with open(PLIK_PORTFELA, "w", encoding="utf-8") as f:
            json.dump(DOMYSLNE_GRUPY, f, indent=4, ensure_ascii=False)
        return DOMYSLNE_GRUPY
    with open(PLIK_PORTFELA, "r", encoding="utf-8") as f:
        return json.load(f)

def zapisz_portfel(grupy: dict):
    with open(PLIK_PORTFELA, "w", encoding="utf-8") as f:
        json.dump(grupy, f, indent=4, ensure_ascii=False)


# ══════════════════════════════════════════════════════════════
#   BATCH FETCH z elastycznym dopasowaniem kluczy
# ══════════════════════════════════════════════════════════════

def _znajdz_w_batch(wyniki_batch: dict, tv_sym: str):
    """
    BUG FIX: TV może zwrócić klucz w innym formacie niż wysłaliśmy.
    Przykład: wysłaliśmy "XETRA:AMD", TV zwraca "XETR:AMD" lub "FWB:AMD".
    Strategia: szukaj po samej nazwie symbolu (część po ':').
    """
    if not wyniki_batch:
        return None

    # 1. Dokładne dopasowanie (uppercase)
    v = wyniki_batch.get(tv_sym) or wyniki_batch.get(tv_sym.upper())
    if v is not None:
        return v

    # 2. Dopasowanie po samym symbolu (ignoruj exchange prefix)
    sam_sym = tv_sym.split(":")[-1].upper() if ":" in tv_sym else tv_sym.upper()
    for klucz, val in wyniki_batch.items():
        if val is None:
            continue
        klucz_sym = klucz.split(":")[-1].upper() if ":" in klucz else klucz.upper()
        if klucz_sym == sam_sym:
            return val

    return None  # Rzeczywiście brak danych


def _batch_fetch(screener: str, symbole: list,
                  interwal_str: str, max_retry: int = 3) -> dict:
    """Jeden batch request dla całego screener."""
    try:
        from tradingview_ta import get_multiple_analysis, Interval
    except ImportError:
        print("  ⚠️  pip install tradingview_ta")
        return {}

    interwal = getattr(Interval, interwal_str, None)
    if not interwal:
        return {}

    for proba in range(max_retry):
        try:
            wyniki = get_multiple_analysis(
                screener=screener,
                interval=interwal,
                symbols=symbole,
            )
            # Odfiltruj None wartości do logowania
            ok   = sum(1 for v in wyniki.values() if v is not None)
            brak = sum(1 for v in wyniki.values() if v is None)
            if brak > 0:
                print(f" (OK:{ok} brak:{brak})", end="", flush=True)
            return wyniki or {}

        except Exception as e:
            err = str(e)
            if "429" in err:
                czekaj = (proba + 1) * 5
                print(f"\n    ⏳ Rate limit → czekam {czekaj}s", end="", flush=True)
                time.sleep(czekaj)
            elif "403" in err:
                return {}
            else:
                if proba == max_retry - 1:
                    print(f"\n    ⚠️  [{screener}]: {err[:55]}", end="")
                time.sleep(2)
    return {}


def pobierz_wszystkie_dane(tickery: list) -> tuple[dict, dict]:
    """
    Batch fetch dla wszystkich tickerów.
    Zwraca (dane_dict, raport_problemow)
    """
    # Grupuj po screener
    grupy: dict[str, dict] = {}  # screener → {tv_sym: ticker}
    bez_mapy = []

    for ticker in tickery:
        info = tv_info(ticker)
        if info:
            scr, tv_sym = info
            if scr not in grupy:
                grupy[scr] = {}
            grupy[scr][tv_sym] = ticker
        else:
            bez_mapy.append(ticker)

    zbior: dict[str, dict] = {t: {} for t in tickery}
    problemy: dict[str, str] = {}

    if bez_mapy:
        for t in bez_mapy:
            problemy[t] = "Brak mapy TV — dodaj ręcznie do MAPA{}"

    for interwal, interwal_tv in TV_INTERWAL.items():
        print(f"\n  📡 [{interwal}]", end="", flush=True)

        for screener, sym_mapa in grupy.items():
            symbole = list(sym_mapa.keys())
            print(f" {screener}({len(symbole)})", end="", flush=True)

            wyniki_batch = _batch_fetch(screener, symbole, interwal_tv)
            time.sleep(1.2)  # przerwa między screenerami

            for tv_sym, ticker in sym_mapa.items():
                analiza = _znajdz_w_batch(wyniki_batch, tv_sym)

                if analiza is not None:
                    zbior[ticker][interwal] = _parsuj_analize(analiza)
                elif ticker not in problemy and interwal == "1d":
                    sam_sym = tv_sym.split(":")[-1]
                    problemy[ticker] = (
                        f"TV nie znalazło '{tv_sym}' — sprawdź symbol na "
                        f"tradingview.com/symbols/{screener}-{sam_sym}"
                    )

    return zbior, problemy


def _parsuj_analize(analiza) -> dict:
    """Parsuje obiekt Analysis z tradingview_ta."""
    ind      = analiza.indicators
    kupno    = analiza.summary["BUY"]
    sprzedaz = analiza.summary["SELL"]
    neutral  = analiza.summary["NEUTRAL"]
    total    = kupno + sprzedaz + neutral
    rec      = analiza.summary["RECOMMENDATION"]
    wynik_tv = round((kupno / total) * 100) if total > 0 else 50

    def ok(v):
        if v is None: return False
        try: return not (isinstance(v, float) and pd.isna(v))
        except: return True

    rsi    = ind.get("RSI")
    macd   = ind.get("MACD.macd")
    sygnal = ind.get("MACD.signal")
    macd_p = ind.get("MACD.macd[1]")
    sig_p  = ind.get("MACD.signal[1]")
    sma50  = ind.get("SMA50")
    sma200 = ind.get("SMA200")
    bb_up  = ind.get("BB.upper")
    bb_low = ind.get("BB.lower")
    atr    = ind.get("ATR")
    cena   = ind.get("close")
    vol    = ind.get("volume")
    vol_sm = ind.get("volume_sma")
    adx    = ind.get("ADX")
    stoch  = ind.get("Stoch.K")
    cci    = ind.get("CCI20")
    mom    = ind.get("Mom")

    sl = tp1 = tp2 = atr_proc = None
    if ok(cena) and ok(atr):
        c, a = float(cena), float(atr)
        sl = round(c - 2.0 * a, 4); tp1 = round(c + 3.0 * a, 4)
        tp2 = round(c + 5.0 * a, 4); atr_proc = round(a / c * 100, 2)

    mc = "Brak"
    if ok(macd) and ok(sygnal):
        m, s = float(macd), float(sygnal)
        if ok(macd_p) and ok(sig_p):
            mp, sp = float(macd_p), float(sig_p)
            if   mp <= sp and m > s: mc = "🔥 Świeży Bullish!"
            elif mp >= sp and m < s: mc = "💀 Świeży Bearish!"
            elif m > s:              mc = "📈 Bullish"
            else:                    mc = "📉 Bearish"
        else:
            mc = "📈 Bullish" if m > s else "📉 Bearish"

    tr = "⚪ Brak danych"
    if ok(sma50) and ok(sma200) and ok(cena):
        s50, s200, c = float(sma50), float(sma200), float(cena)
        if s50 > s200 and c > s200:  tr = "✨ Złoty Krzyż (Hossa)"
        elif s50 < s200:              tr = "☠️ Krzyż Śmierci (Bessa)"
        elif c > s50:                 tr = "📈 Cena nad SMA50"
        else:                         tr = "📉 Cena pod SMA50"

    rw = "⚪ Normalny"
    if ok(vol) and ok(vol_sm) and float(vol_sm) > 0:
        r = float(vol) / float(vol_sm)
        if r > 1.5:   rw = f"🔥 Skokowy ({r:.1f}x)"
        elif r < 0.7: rw = f"😴 Słaby ({r:.1f}x)"

    rb = "⚪ W środku pasma"
    if ok(bb_up) and ok(bb_low) and ok(cena):
        c = float(cena)
        if c >= float(bb_up) * 0.99:   rb = "⚠️ Przy górnej BB (wykupiony)"
        elif c <= float(bb_low) * 1.01: rb = "🟢 Przy dolnej BB (wyprzedany)"

    rrsi = "⚪ Brak danych"
    if ok(rsi):
        r = float(rsi)
        if   r < 30: rrsi = f"🟢 Mocno wyprzedany ({r:.1f})"
        elif r < 40: rrsi = f"🟡 Wyprzedany ({r:.1f})"
        elif r > 75: rrsi = f"🔴 Mocno wykupiony ({r:.1f})"
        elif r > 60: rrsi = f"🟡 Lekko wykupiony ({r:.1f})"
        else:         rrsi = f"⚪ Neutralny ({r:.1f})"

    adx_s = "Brak danych"
    if ok(adx):
        a = float(adx)
        if   a > 40: adx_s = f"💪 Bardzo silny trend (ADX={a:.0f})"
        elif a > 25: adx_s = f"✅ Silny trend (ADX={a:.0f})"
        elif a > 15: adx_s = f"⚪ Słaby trend (ADX={a:.0f})"
        else:         adx_s = f"😴 Brak trendu (ADX={a:.0f})"

    return {
        "cena": round(float(cena), 4) if ok(cena) else None,
        "rekomendacja": rec, "wynik_tv": wynik_tv,
        "kupno": kupno, "sprzedaz": sprzedaz, "neutralny": neutral,
        "rsi": round(float(rsi), 1) if ok(rsi) else None,
        "macd_crossover": mc, "stan_trendu": tr, "sila_trendu": adx_s,
        "radar_rsi": rrsi, "radar_wolumenu": rw, "radar_bb": rb,
        "stoch_k": round(float(stoch), 1) if ok(stoch) else None,
        "cci":     round(float(cci), 1)   if ok(cci)   else None,
        "momentum":round(float(mom), 4)   if ok(mom)   else None,
        "stop_loss": sl, "take_profit_1": tp1,
        "take_profit_2": tp2, "atr_proc": atr_proc,
    }


# ══════════════════════════════════════════════════════════════
#   SYSTEM PUNKTOWY I DECYZJA
# ══════════════════════════════════════════════════════════════

def oblicz_wynik(dane: dict) -> int:
    s = w = 0
    for iv, waga in WAGI.items():
        if iv in dane:
            s += dane[iv].get("wynik_tv", 50) * waga
            w += waga
    return round(s / w) if w else 0

def generuj_decyzje(dane: dict, wynik: int) -> tuple:
    if wynik == 0:
        return "⚫ BRAK DANYCH TV", []

    powody = []
    d1  = dane.get("1d",  {})
    d4h = dane.get("4h",  {})
    dwk = dane.get("1wk", {})

    tw  = dwk.get("stan_trendu", "")
    t1  = d1.get("stan_trendu", "")
    m1  = d1.get("macd_crossover", "")
    r1  = d1.get("rsi")
    adx = d1.get("sila_trendu", "")

    hossa = "Złoty Krzyż" in tw or "Złoty Krzyż" in t1
    bessa = "Krzyż Śmierci" in tw or "Krzyż Śmierci" in t1
    fb    = "Świeży" in m1 and "Bullish" in m1
    sb    = "Świeży" in m1 and "Bearish" in m1

    if hossa:      powody.append("✅ Złoty Krzyż — trend wzrostowy")
    if bessa:      powody.append("❌ Krzyż Śmierci — trend spadkowy")
    if fb:         powody.append("🔥 Świeży Bullish Crossover MACD!")
    if sb:         powody.append("💀 Świeży Bearish Crossover MACD!")
    if r1 and r1 < 35: powody.append(f"🟢 RSI wyprzedany ({r1}) — potencjał odbicia")
    if r1 and r1 > 72: powody.append(f"🔴 RSI wykupiony ({r1}) — ryzyko korekty")
    if "Bardzo silny" in adx: powody.append(f"💪 {adx}")
    if "Skokowy" in d1.get("radar_wolumenu", ""): powody.append(f"📊 {d1['radar_wolumenu']}")
    if "dolnej" in d1.get("radar_bb", ""):  powody.append("🟢 Cena przy dolnej BB")
    if "górnej" in d1.get("radar_bb", ""):  powody.append("⚠️ Cena przy górnej BB")

    if   wynik >= 80 and hossa and fb: dec = "🔥 ALL-IN KUPUJ (Pełna zgodność!)"
    elif wynik >= 75 and hossa:         dec = f"🟢 KUPUJ (Silny trend, {wynik}/100)"
    elif wynik >= 65 and fb:            dec = f"🟢 KUPUJ (Crossover, {wynik}/100)"
    elif wynik >= 60:                   dec = f"🟡 ROZWAŻ KUPNO (TV: {wynik}/100)"
    elif wynik <= 20 and bessa and sb:  dec = "🚨 UCIEKAJ / SPRZEDAJ natychmiast!"
    elif wynik <= 30 and bessa:         dec = f"🔴 SPRZEDAJ (Bessa, {wynik}/100)"
    elif wynik <= 35 and sb:            dec = f"🔴 SPRZEDAJ (Bearish, {wynik}/100)"
    elif wynik <= 40:                   dec = f"🔴 REDUKUJ pozycję (TV: {wynik}/100)"
    else:                               dec = f"⚪ CZEKAJ (TV: {wynik}/100)"

    return dec, powody

def ocena_ryzyka(atr_proc) -> str:
    try:
        if atr_proc is None or pd.isna(atr_proc): return "⚪ Brak danych"
    except: pass
    if not isinstance(atr_proc, (int, float)): return "⚪ Brak danych"
    if   atr_proc < 1.5:  return f"🟢 Niskie ({atr_proc}%)"
    elif atr_proc <= 4.0: return f"🟡 Średnie ({atr_proc}%)"
    elif atr_proc <= 8.0: return f"🔴 Wysokie ({atr_proc}%)"
    else:                  return f"💀 BARDZO WYSOKIE ({atr_proc}%)"


# ══════════════════════════════════════════════════════════════
#   TELEGRAM
# ══════════════════════════════════════════════════════════════

def wyslij_telegram(wiadomosc: str, sciezka: str = None):
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        print("\n⚠️  Brak tokenów Telegram → ustaw TELEGRAM_TOKEN i TELEGRAM_CHAT_ID")
        return
    try:
        if sciezka and os.path.exists(sciezka):
            url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument"
            with open(sciezka, "rb") as f:
                r = requests.post(url, data={
                    "chat_id": TELEGRAM_CHAT_ID,
                    "caption": wiadomosc[:1024],
                    "parse_mode": "HTML",
                }, files={"document": f}, timeout=30)
        else:
            url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
            r = requests.post(url, json={
                "chat_id": TELEGRAM_CHAT_ID,
                "text": wiadomosc, "parse_mode": "HTML",
            }, timeout=15)
        print("\n✅ Telegram wysłano!" if r.status_code == 200 else f"\n❌ Telegram {r.status_code}: {r.text[:100]}")
    except Exception as e:
        print(f"\n❌ Telegram: {e}")

def zbuduj_wiadomosc(wyniki: list, problemy: dict) -> str:
    data = datetime.now().strftime("%d.%m.%Y %H:%M")
    tv_ok   = sum(1 for w in wyniki if w["wynik"] > 0)
    tv_brak = len(problemy)
    msg  = f"🤖 <b>QUANT RADAR v4.1 — {data}</b>\n"
    msg += f"📊 TV: {tv_ok} ✅ | Brak danych: {tv_brak} ⚫\n"
    msg += "━━━━━━━━━━━━━━━━━━━━━━━\n\n"

    aktywne = [w for w in wyniki if any(x in w["decyzja"] for x in ["🔥","🟢","🚨","🔴"])]
    if not aktywne:
        msg += "⚪ Rynek spokojny. Brak wyraźnych sygnałów.\n"
    else:
        for w in aktywne[:8]:
            d1   = w["dane"].get("1d", {})
            info = tv_info(w["ticker"])
            gielda = info[1].split(":")[0] if info else "?"
            msg += f"<b>{w['ticker']}</b> [{gielda}] — {d1.get('cena','?')}\n"
            msg += f"👉 {w['decyzja']}\n"
            msg += f"📊 {w['wynik']}/100 | BUY:{d1.get('kupno',0)} SELL:{d1.get('sprzedaz',0)}\n"
            for p in w.get("powody", [])[:2]: msg += f"   {p}\n"
            if "KUPUJ" in w["decyzja"]:
                msg += f"🛡️ SL:{d1.get('stop_loss','?')} TP:{d1.get('take_profit_1','?')} | {ocena_ryzyka(d1.get('atr_proc'))}\n"
            msg += "\n"

    if problemy:
        msg += f"⚫ <b>Brak danych TV ({len(problemy)}):</b>\n"
        for t, p in list(problemy.items())[:6]:
            msg += f"   {t}: {p[:60]}\n"

    msg += "\n📎 <i>Pełny raport Excel w załączniku.</i>"
    return msg


# ══════════════════════════════════════════════════════════════
#   EXCEL — kolumny auto-fit WE WSZYSTKICH ARKUSZACH
# ══════════════════════════════════════════════════════════════

def _autofit_wszystkie_arkusze(wb):
    """BUG FIX: auto-fit kolumn we wszystkich arkuszach."""
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    val_len = len(str(cell.value)) if cell.value is not None else 0
                    if val_len > max_len:
                        max_len = val_len
                except:
                    pass
            # Ogranicz max szerokość i dodaj padding
            ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def zapisz_excel(wyniki: list, problemy: dict, nazwa: str):
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl import load_workbook

        podsumowanie = []
        for w in wyniki:
            d1  = w["dane"].get("1d",  {})
            dwk = w["dane"].get("1wk", {})
            info = tv_info(w["ticker"])
            powod_bledu = problemy.get(w["ticker"], "")
            podsumowanie.append({
                "Ticker":         w["ticker"],
                "Giełda TV":      info[1].split(":")[0] if info else "?",
                "Symbol TV":      info[1].split(":")[-1] if info else "?",
                "Grupa":          w["grupa"],
                "Cena":           d1.get("cena", "—"),
                "REKOMENDACJA":   w["decyzja"],
                "TV Score":       w["wynik"] if w["wynik"] > 0 else "BRAK",
                "BUY signals":    d1.get("kupno",    "—"),
                "SELL signals":   d1.get("sprzedaz", "—"),
                "Ryzyko ATR":     ocena_ryzyka(d1.get("atr_proc")),
                "Stop Loss":      d1.get("stop_loss",     "—"),
                "Take Profit 1":  d1.get("take_profit_1", "—"),
                "Take Profit 2":  d1.get("take_profit_2", "—"),
                "RSI (1d)":       d1.get("rsi",           "—"),
                "MACD (1d)":      d1.get("macd_crossover","—"),
                "Stoch K (1d)":   d1.get("stoch_k",       "—"),
                "CCI (1d)":       d1.get("cci",           "—"),
                "Momentum":       d1.get("momentum",      "—"),
                "Trend SMA":      d1.get("stan_trendu",   "—"),
                "Siła ADX":       d1.get("sila_trendu",   "—"),
                "Bollinger":      d1.get("radar_bb",      "—"),
                "Wolumen":        d1.get("radar_wolumenu","—"),
                "Trend 1WK":      dwk.get("stan_trendu",  "—"),
                "Powody":         " | ".join(w.get("powody", [])),
                "Błąd TV":        powod_bledu,
            })

        df_sum = pd.DataFrame(podsumowanie)

        # Arkusze interwałowe
        df_int = {}
        for iv in ["1h", "4h", "1d", "1wk"]:
            rows = []
            for w in wyniki:
                d = w["dane"].get(iv)
                if d:
                    rows.append({"Ticker": w["ticker"], "Grupa": w["grupa"],
                                 "Giełda": tv_info(w["ticker"])[1].split(":")[0] if tv_info(w["ticker"]) else "?",
                                 **d})
            if rows:
                df_int[iv] = pd.DataFrame(rows)

        # Arkusz problemów
        if problemy:
            df_problemy = pd.DataFrame([
                {"Ticker": t, "Problem": p,
                 "Jak sprawdzić": f"tradingview.com → szukaj: {t.split('.')[0]}"}
                for t, p in problemy.items()
            ])

        with pd.ExcelWriter(nazwa, engine="openpyxl") as writer:
            df_sum.to_excel(writer, sheet_name="🌟 PODSUMOWANIE", index=False)
            for iv, dfi in df_int.items():
                dfi.to_excel(writer, sheet_name=f"📊 {iv.upper()}", index=False)
            if problemy:
                df_problemy.to_excel(writer, sheet_name="⚫ BRAK DANYCH TV", index=False)

            wb = writer.book

            # ── Kolorowanie PODSUMOWANIA ──────────────────
            ws = wb["🌟 PODSUMOWANIE"]
            KOL = {
                "🔥": ("00B050", "FFFFFF"),  # ciemny zielony
                "🟢": ("C6EFCE", "276221"),  # jasny zielony
                "🔴": ("FFC7CE", "9C0006"),  # jasny czerwony
                "🚨": ("FF4444", "FFFFFF"),  # mocny czerwony
                "🟡": ("FFEB9C", "9C6500"),  # żółty
                "⚪": ("F5F5F5", "666666"),  # szary
                "⚫": ("DDDDDD", "888888"),  # ciemno-szary
            }
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                val = str(row[5].value or "")  # kolumna REKOMENDACJA (index 5)
                for ik, (bg, fg) in KOL.items():
                    if ik in val:
                        fill = PatternFill("solid", fgColor=bg)
                        font_kol = Font(color=fg)
                        for cell in row:
                            cell.fill = fill
                        # Bold + kolor tekstu dla ważnych sygnałów
                        if ik in ("🔥", "🚨", "🟢", "🔴"):
                            row[5].font = Font(bold=True, color=fg)
                        break

            # ── Zamróź górny wiersz we wszystkich arkuszach ──
            for ws_name in wb.sheetnames:
                wb[ws_name].freeze_panes = "A2"

            # ── AUTO-FIT KOLUMN WE WSZYSTKICH ARKUSZACH ──
            # BUG FIX: wcześniej tylko PODSUMOWANIE miało auto-fit
            _autofit_wszystkie_arkusze(wb)

        print(f"\n💾 Zapisano: {nazwa}")

    except Exception as e:
        print(f"\n⚠️  Excel błąd: {e}")
        pd.DataFrame(podsumowanie).to_csv(nazwa.replace(".xlsx", ".csv"), index=False)
        print("   → Zapisano CSV zamiast Excel")


# ══════════════════════════════════════════════════════════════
#   TRYB DIAGNOSTYCZNY — sprawdź pojedynczy ticker
# ══════════════════════════════════════════════════════════════

def sprawdz_ticker(ticker: str):
    """Testuje czy ticker działa z TradingView i pokazuje co zwraca."""
    print(f"\n🔍 Diagnostyka: {ticker}")
    print("=" * 55)

    info = tv_info(ticker)
    if not info:
        print(f"❌ Brak mapy TV dla '{ticker}'")
        print(f"   → Dodaj ręcznie do słownika MAPA{{}} w bot_tv.py")
        return

    screener, tv_sym = info
    exchange, symbol = tv_sym.split(":") if ":" in tv_sym else ("?", tv_sym)
    print(f"   Ticker portfela : {ticker}")
    print(f"   Screener TV     : {screener}")
    print(f"   Giełda TV       : {exchange}")
    print(f"   Symbol TV       : {symbol}")
    print(f"   Pełny klucz     : {tv_sym}")
    print(f"\n   🌐 Sprawdź na: https://tradingview.com/symbols/{exchange}-{symbol}/")
    print()

    try:
        from tradingview_ta import get_multiple_analysis, Interval

        print(f"   Testuję połączenie z TV...", end="", flush=True)
        wyniki = get_multiple_analysis(
            screener=screener,
            interval=Interval.INTERVAL_1_DAY,
            symbols=[tv_sym],
        )
        print(" ✅ Połączono!")
        print(f"\n   Klucze zwrócone przez TV: {list(wyniki.keys())}")

        analiza = _znajdz_w_batch(wyniki, tv_sym)
        if analiza:
            print(f"\n   ✅ Analiza dostępna!")
            print(f"   Rekomendacja : {analiza.summary['RECOMMENDATION']}")
            print(f"   BUY          : {analiza.summary['BUY']}")
            print(f"   SELL         : {analiza.summary['SELL']}")
            print(f"   Cena         : {analiza.indicators.get('close')}")
            print(f"   RSI          : {analiza.indicators.get('RSI')}")
        else:
            print(f"\n   ❌ Symbol '{tv_sym}' nie znaleziony w TV")
            print(f"   → Możliwe przyczyny:")
            print(f"     1. Błędna nazwa exchange ({exchange})")
            print(f"     2. Symbol niedostępny na tej giełdzie w TV")
            print(f"     3. Spółka wycofana/zawieszona")
            print(f"   → Idź na tradingview.com i szukaj: '{symbol}'")
            print(f"     Sprawdź jaki exchange TV pokazuje w URL")

    except Exception as e:
        err = str(e)
        if "403" in err:
            print(f" ⚠️  Serwer blokuje połączenie (403)")
            print(f"   → Ten test działa tylko na GitHub Actions lub lokalnie")
            print(f"   → Na GitHub Actions ticker powinien działać")
        else:
            print(f" ❌ Błąd: {err}")


# ══════════════════════════════════════════════════════════════
#   MAIN
# ══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Quant Radar Bot v4.1")
    parser.add_argument("--dodaj",    type=str, help="Ticker do dodania")
    parser.add_argument("--usun",     type=str, help="Ticker do usunięcia")
    parser.add_argument("--lista",    action="store_true", help="Pokaż portfel")
    parser.add_argument("--sprawdz",  type=str, help="Testuj ticker (np. --sprawdz KGHM.WA)")
    parser.add_argument("--grupa",    type=str, default="💼 MÓJ PORTFEL")
    args = parser.parse_args()

    # ── Tryb diagnostyczny ────────────────────────────────
    if args.sprawdz:
        sprawdz_ticker(args.sprawdz.upper())
        return

    grupy = zaladuj_portfel()

    if args.lista:
        print("\n📋 Portfel i mapowania TV:")
        tv_ok = tv_brak = 0
        for nazwa_g, lista in grupy.items():
            print(f"\n  {nazwa_g}:")
            for t in lista:
                info = tv_info(t)
                if info:
                    print(f"    ✅ {t:<15} → {info[1]} (screener: {info[0]})")
                    tv_ok += 1
                else:
                    print(f"    ❌ {t:<15} → BRAK MAPY TV")
                    tv_brak += 1
        print(f"\n  Suma: ✅ {tv_ok} zmapowanych | ❌ {tv_brak} bez mapy")
        return

    if args.dodaj:
        sym = args.dodaj.upper()
        if args.grupa not in grupy: grupy[args.grupa] = []
        if sym not in grupy[args.grupa]:
            grupy[args.grupa].append(sym)
            zapisz_portfel(grupy)
        info = tv_info(sym)
        if info:
            print(f"✅ Dodano {sym} → TV: {info[1]} (screener: {info[0]})")
        else:
            print(f"⚠️  Dodano {sym} ale brak mapy TV")
            print(f"   Uruchom: python bot_tv.py --sprawdz {sym}")
        return

    if args.usun:
        sym = args.usun.upper()
        for g in grupy.values():
            if sym in g: g.remove(sym)
        zapisz_portfel(grupy)
        print(f"🗑️  Usunięto {sym}")
        return

    # ── Pełne skanowanie ─────────────────────────────────
    start = time.time()
    print("╔══════════════════════════════════════════════╗")
    print("║  🤖 QUANT RADAR v4.1 — Fix Edition          ║")
    print(f"║  {datetime.now().strftime('%Y-%m-%d %H:%M'):<44}║")
    print("╚══════════════════════════════════════════════╝")

    wszystkie = {}
    for nazwa_g, lista in grupy.items():
        for t in lista:
            if t not in wszystkie:
                wszystkie[t] = nazwa_g

    print(f"\n📋 Portfel: {len(wszystkie)} instrumentów | Źródło: wyłącznie TradingView\n")

    zbior, problemy = pobierz_wszystkie_dane(list(wszystkie.keys()))

    wyniki_finalne = []
    for ticker, grupa in wszystkie.items():
        dane    = zbior.get(ticker, {})
        wynik   = oblicz_wynik(dane)
        decyzja, powody = generuj_decyzje(dane, wynik)
        wyniki_finalne.append({
            "ticker": ticker, "grupa": grupa, "wynik": wynik,
            "decyzja": decyzja, "powody": powody, "dane": dane,
        })

    wyniki_finalne.sort(key=lambda x: x["wynik"], reverse=True)

    czas = time.time() - start
    tv_ok   = sum(1 for w in wyniki_finalne if w["wynik"] > 0)
    tv_brak = sum(1 for w in wyniki_finalne if w["wynik"] == 0)

    print(f"\n{'='*62}")
    print(f"  🏆 RANKING  (czas: {czas:.0f}s | TV✅:{tv_ok} | Brak⚫:{tv_brak})")
    print(f"{'='*62}")
    for w in wyniki_finalne:
        ikona = "✅" if w["wynik"] > 0 else "⚫"
        print(f"  {ikona} {w['ticker']:<12} {w['wynik']:>3}/100  {w['decyzja'][:45]}")

    if problemy:
        print(f"\n{'='*62}")
        print(f"  ⚫ BRAK DANYCH TV — przyczyny:")
        print(f"{'='*62}")
        for t, p in problemy.items():
            print(f"  {t:<12} → {p[:58]}")
        print(f"\n  💡 Sprawdź: python bot_tv.py --sprawdz <TICKER>")

    data_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    nazwa    = f"Raport_QuantRadar_{data_str}.xlsx"
    zapisz_excel(wyniki_finalne, problemy, nazwa)
    wyslij_telegram(zbuduj_wiadomosc(wyniki_finalne, problemy), sciezka=nazwa)


if __name__ == "__main__":
    main()